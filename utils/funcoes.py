"""
Funções utilitárias responsáveis pela coleta, leitura de arquivos,
persistência em banco de dados e processamento de páginas.
"""


import os
import re

import openpyxl
import pyodbc
from openpyxl import load_workbook
import requests

from utils.config import (
    CAMINHO_METADADOS,
    STR_CONEXAO_ACCESS,
    CAMINHO_ALVOS,
)
from utils.instrucoes_sql import (
    SELECAO_IDS_BANCO_DE_DADOS,
    SELECAO_IDS_METADADOS
)
from utils.classes.Excecoes import (
    ErroPastaDeAlvosVazia,
    ErroNenhumAlvoPassado,
    ErroNenhumArquivoValidoEmAlvos,
    AlvosJaConstamNoBancoDeDados
)
from utils.persistence.access import selecionar_ids, excluir_do_banco, excluir_dos_metadados
from utils.transformation.data_manipulation import transformar_em_url


def coletar_alvos(conexao: pyodbc.Connection) -> list[str]:
    """Coleta os identificadores válidos presentes nos arquivos de alvos."""
    arquivos_na_pasta_alvo: list[str] = os.listdir(CAMINHO_ALVOS)
    if not arquivos_na_pasta_alvo:
        raise ErroPastaDeAlvosVazia

    extensoes_validas: list[str] = ['xls', 'xlsx']
    arquivos_alvo: list[str] = []
    for nome_arquivo in arquivos_na_pasta_alvo:
        if nome_arquivo.split('.')[-1] in extensoes_validas:
            arquivos_alvo.append(nome_arquivo)

    if not arquivos_alvo:
        raise ErroNenhumArquivoValidoEmAlvos

    alvos: list[str] = []
    for nome_arquivo in arquivos_alvo:
        extensao = nome_arquivo.split('.')[-1]
        caminho_ao_arquivo = f'{CAMINHO_ALVOS}/{nome_arquivo}'
        alvos_encontrados: list[str] = []
        if extensao == 'xls':
            alvos_encontrados += extrair_de_xls(caminho_ao_arquivo)
        elif extensao == 'xlsx':
            alvos_encontrados += extrair_de_xlsx(caminho_ao_arquivo)

        for alvo in alvos_encontrados:
            if alvo not in alvos:
                alvos.append(alvo)

    if not alvos:
        raise ErroNenhumAlvoPassado

    if not isinstance(conexao, str):
        alvos = verificar_pendencia(alvos, conexao)

    if not alvos:
        raise AlvosJaConstamNoBancoDeDados

    return alvos


def extrair_de_xlsx(nome_arquivo: str) -> list[str]:
    """Extrai identificadores de páginas de um arquivo '.xlsx'."""
    extrato: list[str] = []
    livro_alvos: openpyxl.Workbook = load_workbook(nome_arquivo)
    folha_alvos = livro_alvos.active
    for row in folha_alvos.iter_rows(1, folha_alvos.max_row):
        for col in range(len(row)):
            conteudo_celula: str = str(row[col].value)
            if e_id_pagina(conteudo_celula):
                extrato.append(conteudo_celula)

    livro_alvos.close()
    return extrato


def extrair_de_xls(nome_arquivo: str) -> list[str]:
    """Extrai identificadores de páginas de um arquivo '.xls'."""
    extrato: list[str] = []
    with open(nome_arquivo, 'r') as dados_arquivo:
        dados_xls: list[str] = ''.join(dados_arquivo.readlines()).replace("</td>", "").split("<td>")
        for dado_xls in dados_xls:
            if e_id_pagina(dado_xls):
                extrato.append(dado_xls)
    return extrato


def e_id_pagina(possivel_id: str) -> bool:
    """Retorna se o texto corresponde a um identificador de página."""
    return bool(re.match('^[0-9]{8}$', possivel_id))


def verificar_pendencia(paginas: list[str], conexao: pyodbc.Connection) -> list[str]:
    """Remove inconsistências e retorna as páginas ainda pendentes."""
    paginas_pendentes: list[str] = []
    paginas_para_excluir_do_banco: list[str] = []
    paginas_para_excluir_do_metadados: dict[str, list[str]] = {}

    arquivos_metadados = os.listdir(CAMINHO_METADADOS)
    for nome_arquivo in arquivos_metadados:
        extensao = nome_arquivo.split('.')[-1]
        if extensao != "accdb":
            arquivos_metadados.remove(nome_arquivo)

    paginas_no_banco = selecionar_ids(conexao, SELECAO_IDS_BANCO_DE_DADOS)
    metadados: dict[str, str] = {}

    if arquivos_metadados:
        for nome_arquivo in arquivos_metadados:
            conexao_metadados = pyodbc.connect(STR_CONEXAO_ACCESS.format(CAMINHO_METADADOS + nome_arquivo))
            paginas_nos_metadados = selecionar_ids(conexao_metadados, SELECAO_IDS_METADADOS)

            for pagina in paginas_nos_metadados:
                metadados[pagina.strip()] = nome_arquivo

            conexao_metadados.close()

    for pagina in paginas:
        pagina_esta_no_banco = pagina in paginas_no_banco
        pagina_esta_nos_metadados = pagina in metadados
        if (not pagina_esta_no_banco) and (not pagina_esta_nos_metadados):
            paginas_pendentes.append(pagina)

        else:
            if pagina_esta_no_banco and (not pagina_esta_nos_metadados):
                paginas_para_excluir_do_banco.append(pagina)

            elif pagina_esta_nos_metadados and (not pagina_esta_no_banco):
                try:
                    paginas_para_excluir_do_metadados[metadados[pagina]].append(pagina)
                except KeyError:
                    paginas_para_excluir_do_metadados[metadados[pagina]] = [pagina]

    excluir_do_banco(paginas_para_excluir_do_banco, conexao)
    excluir_dos_metadados(paginas_para_excluir_do_metadados)

    return paginas_pendentes


def raspar_conteudo(_id: str, _sessao: requests.Session, _ck: dict[str, str]) -> str:
    """Obtém o conteúdo HTML de uma página."""
    url = transformar_em_url(_id)
    return _sessao.get(url, cookies=_ck).text
